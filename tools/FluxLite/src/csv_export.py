from __future__ import annotations

import csv
import os
from typing import List

from . import config


HEADERS: List[str] = ["DeviceID", "Pass/Fail", "DateTime", "Tester", "BodyWeightN", "ModelID"]


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(path)
    if parent and not os.path.isdir(parent):
        os.makedirs(parent, exist_ok=True)


def _format_excel_safe_text(value: str) -> str:
    """Return a CSV-safe string that keeps leading zeros in Excel.

    Excel strips leading zeros for numeric-looking fields in CSV. The common
    workaround is to write a formula-like string ="000123" so Excel keeps
    it as text with leading zeros preserved. IDEs or text viewers will
    display the literal, which is acceptable for this use case.
    """
    s = (value or "").strip()
    if not s:
        return ""
    # If it looks numeric and starts with 0, force Excel text preservation
    is_numeric_like = s.replace(".", "", 1).isdigit()
    if s.startswith("0") and is_numeric_like:
        return f'="{s}"'
    return s


def _append_csv_row(out_path: str, row: list[str]) -> None:
    ensure_parent_dir(out_path)
    file_exists = os.path.exists(out_path)
    # Use UTF-8 with BOM so Excel detects UTF-8, CRLF endings, and quote all
    with open(out_path, mode=("a" if file_exists else "w"), newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        if not file_exists:
            writer.writerow(HEADERS)
        writer.writerow(row)


def _append_xlsx_row(out_path: str, row: list[str]) -> None:
    ensure_parent_dir(out_path)
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as _:
        # Fallback to CSV if openpyxl is unavailable
        csv_fallback = os.path.splitext(out_path)[0] + ".csv"
        _append_csv_row(csv_fallback, row)
        return

    if os.path.exists(out_path):
        try:
            wb = load_workbook(out_path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()
    ws = wb.active
    # If sheet is empty, write headers first
    if ws.max_row == 1 and all((ws.cell(row=1, column=i+1).value is None) for i in range(len(HEADERS))):
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.number_format = "@" if header in ("DeviceID", "ModelID") else cell.number_format
    # Append row, set text format for ID fields to preserve leading zeros
    next_row = ws.max_row + 1
    for col, value in enumerate(row, start=1):
        cell = ws.cell(row=next_row, column=col)
        cell.value = value
        if col in (1, 6):  # DeviceID, ModelID
            cell.number_format = "@"
    wb.save(out_path)


def append_summary_row(device_id: str, pass_fail: str, date_text: str, tester: str, body_weight_n: float, model_id: str, path: str | None = None) -> str:
    """Append a summary row to CSV or XLSX depending on the file extension.

    - If `path` (or config default) ends with .xlsx, write with openpyxl.
    - Otherwise, write CSV with UTF-8 BOM, CRLF, and quoting to help Excel.
    - Creates the file with headers if it doesn't exist.
    """
    out_path = (path or getattr(config, "CSV_EXPORT_PATH", "") or "").strip()
    if not out_path:
        out_path = os.path.join(os.getcwd(), "LiveTesting_Summary.csv")

    # Build row values; coerce numeric fields appropriately
    # For CSV, massage DeviceID and ModelID to preserve leading zeros in Excel
    device_csv = _format_excel_safe_text(str(device_id))
    model_csv = _format_excel_safe_text(str(model_id))
    row_csv = [device_csv, str(pass_fail or ""), str(date_text or ""), str(tester or ""), f"{float(body_weight_n):.6f}", model_csv]
    row_xlsx = [str(device_id or ""), str(pass_fail or ""), str(date_text or ""), str(tester or ""), float(body_weight_n), str(model_id or "")]

    _, ext = os.path.splitext(out_path.lower())
    if ext == ".xlsx":
        _append_xlsx_row(out_path, row_xlsx)
    else:
        # Default to CSV for unknown or missing extensions
        if not ext:
            out_path = out_path + ".csv"
        _append_csv_row(out_path, row_csv)
    return out_path


def append_summary_row_csv(device_id: str, pass_fail: str, date_text: str, tester: str, body_weight_n: float, model_id: str, path: str | None = None) -> str:
    """Backward-compatible CSV-only entry point.

    Prefer `append_summary_row`, which also supports .xlsx.
    """
    if path and path.lower().endswith(".xlsx"):
        return append_summary_row(device_id, pass_fail, date_text, tester, body_weight_n, model_id, path)
    # Force CSV behavior
    chosen = path or getattr(config, "CSV_EXPORT_PATH", "") or os.path.join(os.getcwd(), "LiveTesting_Summary.csv")
    if not chosen.lower().endswith(".csv"):
        chosen = os.path.splitext(chosen)[0] + ".csv"
    return append_summary_row(device_id, pass_fail, date_text, tester, body_weight_n, model_id, chosen)


